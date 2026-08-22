"""
Команда "Заказ" из модуля "Работа".

Интерактивное заполнение Excel-шаблона голосом:
1. Создает файл из шаблона D:/AiA/Work/заказ.xlsx
2. Загружает таблицу подмен с Листа 2 (столбец A — фразы Vosk, столбец B — подстановка)
3. Последовательно озвучивает заголовки из первой строки Листа 1
4. Ждет ответа от Android (Vosk), подменяет по словарю, записывает в строку 2
5. По окончании сохраняет файл и озвучивает "Запись окончена"

Версия 3.4: добавлена подмена распознанных фраз по таблице соответствий (Лист 2).

Протокол WebSocket:
 Сервер → Android: {"type": "zakaz_prompt", "text": "...", "col": N, "total": M, "has_tts": true}
 Android → Сервер: {"type": "command", "module": "work", "command": "zakaz", "text": "..."}
 Android → Сервер: {"type": "command", "module": "work", "command": "zakaz", "text": "__CANCEL__"}
 Сервер → Android: {"type": "tts_started"}
 Сервер → Android: {"type": "tts_ended"}
 Сервер → Android: {"type": "zakaz_done", "message": "..."}
"""
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Any

from openpyxl import load_workbook

BASE_DIR = Path("D:/AiA") if Path("D:/").exists() else Path.home() / "AiA"
WORK_DIR = BASE_DIR / "Work"
TEMPLATE_FILE = WORK_DIR / "заказ.xlsx"

_sessions: Dict[str, Dict[str, Any]] = {}


async def execute(ws_manager, client_ip: str, data: Dict[str, Any],
                  ollama_chat, ollama_model: str, tts_engine, logger):
    """Главная точка входа команды Заказ."""
    raw_text = data.get("text", "").strip()
    converted_text = data.get("converted_text", raw_text).strip()

    if converted_text == "__CANCEL__":
        await _cancel_zakaz(ws_manager, client_ip, logger)
        return

    if client_ip in _sessions:
        await _process_answer(ws_manager, client_ip, converted_text, raw_text, tts_engine, logger)
        return

    await _start_zakaz(ws_manager, client_ip, tts_engine, logger)


async def _start_zakaz(ws_manager, client_ip, tts_engine, logger):
    """Создает файл из шаблона, загружает словарь подмен, отправляет первый вопрос."""
    now = datetime.now()
    mm_yyyy = now.strftime("%m_%Y")
    month_dir = WORK_DIR / mm_yyyy
    month_dir.mkdir(parents=True, exist_ok=True)

    filename = now.strftime("заказ_%m_%d_%H_%M.xlsx")
    dest_file = month_dir / filename

    if not TEMPLATE_FILE.exists():
        msg = f"Шаблон {TEMPLATE_FILE} не найден. Создайте файл-шаблон."
        logger.error(f"[zakaz] {msg}")
        await _send_error(ws_manager, client_ip, msg, tts_engine)
        return

    try:
        shutil.copy2(TEMPLATE_FILE, dest_file)
        logger.info(f"[zakaz] Создан файл: {dest_file}")
    except Exception as e:
        msg = f"Ошибка копирования шаблона: {e}"
        logger.error(f"[zakaz] {msg}")
        await _send_error(ws_manager, client_ip, msg, tts_engine)
        return

    try:
        wb = load_workbook(dest_file)
        ws = wb.active
    except Exception as e:
        msg = f"Ошибка открытия Excel: {e}"
        logger.error(f"[zakaz] {msg}")
        await _send_error(ws_manager, client_ip, msg, tts_engine)
        return

    total_cols = 0
    for col in range(1, 100):
        val = ws.cell(row=1, column=col).value
        if val is None or str(val).strip() == "":
            break
        total_cols += 1

    if total_cols == 0:
        msg = "Шаблон пустой — нет заголовков в первой строке"
        logger.error(f"[zakaz] {msg}")
        await _send_error(ws_manager, client_ip, msg, tts_engine)
        wb.close()
        return

    # === Загрузка таблицы подмен с Листа 2 ===
    mapping: Dict[str, str] = {}
    if len(wb.worksheets) > 1:
        mapping_ws = wb.worksheets[1]
        for row in range(1, 10000):
            key = mapping_ws.cell(row=row, column=1).value
            val = mapping_ws.cell(row=row, column=2).value
            if key is None or str(key).strip() == "":
                break
            if val is not None:
                mapping[str(key).strip().lower()] = str(val).strip()
        logger.info(f"[zakaz] Загружено {len(mapping)} правил подмены с Листа 2")
    else:
        logger.info("[zakaz] Лист 2 не найден — подмена отключена")
    # ==========================================

    _sessions[client_ip] = {
        "wb": wb,
        "ws": ws,
        "current_col": 1,
        "total_cols": total_cols,
        "file_path": dest_file,
        "mapping": mapping,
    }

    logger.info(f"[zakaz] Начат заказ, колонок для заполнения: {total_cols}")
    await _send_next_prompt(ws_manager, client_ip, tts_engine, logger)


async def _process_answer(ws_manager, client_ip, text, raw_text, tts_engine, logger):
    """Записывает ответ пользователя (с подменой по словарю) и отправляет следующий вопрос."""
    session = _sessions.get(client_ip)
    if not session:
        logger.warning(f"[zakaz] Ответ от {client_ip}, но сессии нет")
        return

    ws = session["ws"]
    col = session["current_col"]
    mapping = session.get("mapping", {})

    # === Подмена по таблице соответствий ===
    lookup_key = text.lower().strip()
    if lookup_key in mapping:
        text_to_write = mapping[lookup_key]
        logger.info(f"[zakaz] Подмена: '{text}' → '{text_to_write}'")
    else:
        text_to_write = text
    # =======================================

    ws.cell(row=2, column=col, value=text_to_write)

    if text != raw_text:
        logger.info(
            f"[zakaz] Колонка {col}/{session['total_cols']}: "
            f"записано '{text_to_write}' (исходный: '{raw_text}', преобразованный: '{text}')"
        )
    else:
        logger.info(f"[zakaz] Колонка {col}/{session['total_cols']}: записано '{text_to_write}'")

    session["current_col"] += 1

    if session["current_col"] > session["total_cols"]:
        await _finish_zakaz(ws_manager, client_ip, tts_engine, logger)
    else:
        await _send_next_prompt(ws_manager, client_ip, tts_engine, logger)


async def _send_next_prompt(ws_manager, client_ip, tts_engine, logger):
    """Отправляет Android очередной вопрос (zakaz_prompt) и озвучивает."""
    session = _sessions.get(client_ip)
    if not session:
        return

    ws = session["ws"]
    col = session["current_col"]
    header = ws.cell(row=1, column=col).value
    header_str = str(header).strip() if header else ""

    logger.info(f"[zakaz] Вопрос {col}/{session['total_cols']}: {header_str}")

    await ws_manager.send_to(client_ip, {
        "type": "zakaz_prompt",
        "text": header_str,
        "col": col,
        "total": session["total_cols"],
        "has_tts": True,
        "module": "work",
        "command": "zakaz"
    })

    if tts_engine:
        tts_engine.speak(header_str)


async def _finish_zakaz(ws_manager, client_ip, tts_engine, logger):
    """Сохраняет файл, отправляет zakaz_done, озвучивает завершение."""
    session = _sessions.pop(client_ip, None)
    if not session:
        return

    file_path = session["file_path"]
    wb = session["wb"]

    try:
        wb.save(file_path)
        wb.close()
        msg = f'Файл "{file_path.name}" сохранен'
        logger.info(f"[zakaz] {msg}")

        await ws_manager.send_to(client_ip, {
            "type": "zakaz_done",
            "message": msg,
            "module": "work",
            "command": "zakaz"
        })

        if tts_engine:
            tts_engine.speak("Запись окончена")

    except Exception as e:
        logger.error(f"[zakaz] Ошибка сохранения: {e}")
        await ws_manager.send_to(client_ip, {
            "type": "error",
            "message": f"Ошибка сохранения: {e}",
            "module": "work",
            "command": "zakaz"
        })


async def _cancel_zakaz(ws_manager, client_ip, logger):
    """Отменяет заказ по запросу клиента."""
    session = _sessions.pop(client_ip, None)
    if session:
        try:
            session["wb"].close()
        except Exception:
            pass
        logger.info(f"[zakaz] Заказ отменен клиентом {client_ip}")
    else:
        logger.info(f"[zakaz] Отмена от {client_ip}, но сессии не было")

    await ws_manager.send_to(client_ip, {
        "type": "zakaz_done",
        "message": "Заказ отменен",
        "module": "work",
        "command": "zakaz"
    })


async def _send_error(ws_manager, client_ip, msg, tts_engine):
    """Отправляет ошибку клиенту и озвучивает."""
    await ws_manager.send_to(client_ip, {
        "type": "error",
        "message": msg,
        "module": "work",
        "command": "zakaz"
    })
    if tts_engine:
        tts_engine.speak(msg)