"""
Команда "Перескажи" из модуля "Команды пользователя".

Мгновенная команда (без Vosk).
Ищет файл "перескажи.*" в D:/AiA/Command_user/ с расширениями:
.doc, .docx, .pdf, .xls, .xlsx.

- 0 файлов → сообщение "Нет файлов для пересказа"
- >1 файла → сообщение "Оставьте только один файл с именем перескажи"
- 1 файл   → читает содержимое, озвучивает через TTS, пишет в логи

Поддерживает остановку по кнопке "Стоп" (через tts_engine.stop()).
"""
import asyncio
import os
from pathlib import Path
from typing import Dict, Any

BASE_DIR = Path("D:/AiA") if os.path.exists("D:/") else Path.home() / "AiA"
SEARCH_DIR = BASE_DIR / "Command_user"

# Расширения файлов для поиска
ALLOWED_EXTENSIONS = {".doc", ".docx", ".pdf", ".xls", ".xlsx"}


def _find_resume_files() -> list:
    """Ищет файлы "перескажи.*" в SEARCH_DIR."""
    if not SEARCH_DIR.exists():
        return []
    files = []
    for ext in ALLOWED_EXTENSIONS:
        for f in SEARCH_DIR.glob(f"перескажи{ext}"):
            if f.is_file():
                files.append(f)
    return files


def _read_docx(path: Path) -> str:
    try:
        from docx import Document
        doc = Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except ImportError:
        return "[Ошибка: библиотека python-docx не установлена]"
    except Exception as e:
        return f"[Ошибка чтения .docx: {e}]"


def _read_doc(path: Path) -> str:
    try:
        import textract
        return textract.process(str(path)).decode("utf-8", errors="ignore")
    except ImportError:
        return "[Ошибка: библиотека textract не установлена]"
    except Exception as e:
        return f"[Ошибка чтения .doc: {e}]"


def _read_pdf(path: Path) -> str:
    try:
        import PyPDF2
        text = []
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
        return "\n".join(text)
    except ImportError:
        return "[Ошибка: библиотека PyPDF2 не установлена]"
    except Exception as e:
        return f"[Ошибка чтения PDF: {e}]"


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        text = []
        for sheet in wb.worksheets:
            text.append(f"--- Лист: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text.append(row_text)
        return "\n".join(text)
    except ImportError:
        return "[Ошибка: библиотека openpyxl не установлена]"
    except Exception as e:
        return f"[Ошибка чтения .xlsx: {e}]"


def _read_xls(path: Path) -> str:
    try:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        text = []
        for sheet in wb.sheets():
            text.append(f"--- Лист: {sheet.name} ---")
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                row_text = " | ".join(str(cell) for cell in row if cell != "")
                if row_text.strip():
                    text.append(row_text)
        return "\n".join(text)
    except ImportError:
        return "[Ошибка: библиотека xlrd не установлена]"
    except Exception as e:
        return f"[Ошибка чтения .xls: {e}]"


def _read_file(path: Path) -> str:
    """Читает содержимое файла в зависимости от расширения."""
    ext = path.suffix.lower()
    if ext == ".docx":
        return _read_docx(path)
    elif ext == ".doc":
        return _read_doc(path)
    elif ext == ".pdf":
        return _read_pdf(path)
    elif ext == ".xlsx":
        return _read_xlsx(path)
    elif ext == ".xls":
        return _read_xls(path)
    else:
        return f"[Неподдерживаемый формат: {ext}]"


async def execute(ws_manager, client_ip: str, data: Dict[str, Any],
                  ollama_chat, ollama_model: str, tts_engine, logger):
    """
    Выполняет команду "Перескажи".
    Мгновенная команда — не требует голосового ввода.
    """
    try:
        logger.info("[resume] Запуск команды Перескажи...")

        files = _find_resume_files()
        logger.info(f"[resume] Найдено файлов: {len(files)}")

        if len(files) == 0:
            msg = "Нет файлов для пересказа"
            logger.warning(f"[resume] {msg}")
            await ws_manager.send_to(client_ip, {
                "type": "response",
                "text": msg,
                "module": "command_user",
                "command": "resume"
            })
            if tts_engine:
                tts_engine.speak(msg)
            return

        if len(files) > 1:
            msg = "Оставьте только один файл с именем перескажи"
            logger.warning(f"[resume] {msg}. Найдено: {[f.name for f in files]}")
            await ws_manager.send_to(client_ip, {
                "type": "response",
                "text": msg,
                "module": "command_user",
                "command": "resume"
            })
            if tts_engine:
                tts_engine.speak(msg)
            return

        # Ровно один файл — читаем и озвучиваем
        file_path = files[0]
        logger.info(f"[resume] Чтение файла: {file_path}")

        await ws_manager.send_to(client_ip, {
            "type": "status",
            "message": f"Чтение файла {file_path.name}...",
            "module": "command_user",
            "command": "resume"
        })

        content = _read_file(file_path)
        logger.info(f"[resume] Прочитано {len(content)} символов")

        if content.startswith("[Ошибка:"):
            logger.error(f"[resume] {content}")
            await ws_manager.send_to(client_ip, {
                "type": "error",
                "message": content,
                "module": "command_user",
                "command": "resume"
            })
            if tts_engine:
                tts_engine.speak(content)
            return

        # Отправляем клиенту для отображения
        await ws_manager.send_to(client_ip, {
            "type": "response",
            "text": content[:500] + ("..." if len(content) > 500 else ""),
            "module": "command_user",
            "command": "resume"
        })

        # Озвучиваем
        if tts_engine:
            logger.info("[resume] Запуск озвучивания...")
            tts_engine.speak(content)
            logger.info("[resume] Озвучивание запущено")
        else:
            logger.warning("[resume] TTS не инициализирован")

    except Exception as e:
        logger.error(f"[resume] Ошибка: {e}", exc_info=True)
        await ws_manager.send_to(client_ip, {
            "type": "error",
            "message": f"Ошибка команды Перескажи: {e}",
            "module": "command_user",
            "command": "resume"
        })