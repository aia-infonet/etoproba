"""
Команда "Лунный день" из модуля "Команды пользователя".

УЛУЧШЕННЫЙ АЛГОРИТМ вычисления лунного дня.

Проблема старого алгоритма:
    Синодический месяц варьируется от ~29.27 до ~29.83 дней,
    но старый код всегда делил на 30 равных частей.
    Это давало погрешность до ±10 часов на границах месяца.

Решение (ВАРИАНТ 1 — Адаптивное деление, по умолчанию):
    1. Вычисляем ТОЧНУЮ длительность текущего синодического месяца
       (интервал между предыдущим и следующим новолунием).
    2. Если длительность < 29.5 дней → 29 лунных дней,
       иначе → 30 лунных дней.
    3. Делим реальный синодический месяц на N равных частей.
    4. Каждая часть ≈ 24.3 часа (в 29-дневном) или ≈ 23.6 часа (в 30-дневном).

АЛЬТЕРНАТИВА (ВАРИАНТ 2 — Титхи):
    Лунный день = интервал прироста разности долгот Луны–Солнце на 12°.
    Включается константой MOON_DAY_METHOD = "tithi".
    Более точный астрономически, но дни имеют разную длительность (19–26 ч).
"""

import ephem
from datetime import datetime, timezone, timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
import os

BASE_DIR = Path("D:/AiA") if os.path.exists("D:/") else Path.home() / "AiA"
EXCEL_FILE = BASE_DIR / "Command_user" / "moon-day.xlsx"

# ═══════════════════════════════════════════════════════════
# ВЫБОР МЕТОДА: "adaptive" (по умолчанию) или "tithi"
# ═══════════════════════════════════════════════════════════
MOON_DAY_METHOD = "adaptive"   # "adaptive" | "tithi"

# Порог для адаптивного метода: < 29.5 дней → 29 дней, иначе → 30
SYNODIC_THRESHOLD_DAYS = 29.5

# ═══════════════════════════════════════════════════════════
# ШАБЛОННЫЕ ОПИСАНИЯ (если Excel-файл отсутствует)
# ═══════════════════════════════════════════════════════════

DEFAULT_DESCRIPTIONS = [
    "Новолуние. Начало нового лунного цикла. Время для постановки целей и задумок.",
    "Растущая Луна. Символ роста и развития. Подходит для начала новых дел.",
    "Растущая Луна. Активность и энергия нарастают. Хороший день для общения.",
    "Растущая Луна. Время собирать информацию и учиться новому.",
    "Растущая Луна. День силы и решимости. Подходит для важных решений.",
    "Растущая Луна. Гармония и баланс. Благоприятный день для творчества.",
    "Растущая Луна. Энергия достижения пиков. Время для активных действий.",
    "Первая четверть. Половина пути к полнолунию. Проверьте планы.",
    "Растущая Луна. Интуиция усиливается. Доверяйте внутреннему голосу.",
    "Растущая Луна. День изобилия. Подходит для финансовых операций.",
    "Растущая Луна. Эмоции на высоте. Будьте внимательны к близким.",
    "Растущая Луна. Предполнолуние. Завершайте начатые дела.",
    "Полнолуние. Пик лунного цикла. Время ясности и завершения.",
    "Убывающая Луна. Начало фазы освобождения. Избавляйтесь от лишнего.",
    "Убывающая Луна. Время анализа и размышлений. Подведите итоги.",
    "Убывающая Луна. Энергия спадает. Отдыхайте и восстанавливайте силы.",
    "Убывающая Луна. День очищения. Подходит для уборки и порядка.",
    "Убывающая Луна. Время прощения и отпускания обид.",
    "Убывающая Луна. Последняя четверть. Проверьте, что осталось несделанным.",
    "Убывающая Луна. Глубокий анализ. Время для внутренней работы.",
    "Убывающая Луна. Подготовка к новому циклу. Планируйте будущее.",
    "Убывающая Луна. Тихий день. Медитация и самопознание.",
    "Убывающая Луна. Освобождение от старого. Принимайте изменения.",
    "Убывающая Луна. День благодарности. Благодарите за пройденный путь.",
    "Убывающая Луна. Завершение цикла. Подводите окончательные итоги.",
    "Убывающая Луна. Предноволунье. Отпустите всё ненужное.",
    "Убывающая Луна. Последний рывок перед покоем. Завершайте мелочи.",
    "Убывающая Луна. День тишины и покоя. Наблюдайте за собой.",
    "Убывающая Луна. Подготовка к новолунию. Очищайте мысли и пространство.",
    "Убывающая Луна. Финальный день цикла. Отдыхайте и набирайтесь сил."
]


def ensure_excel_exists():
    """Создаёт Excel-файл с шаблонными описаниями, если его нет."""
    if EXCEL_FILE.exists():
        return
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Лунные дни"
    for i, desc in enumerate(DEFAULT_DESCRIPTIONS, start=1):
        ws.cell(row=i, column=1, value=desc)
    wb.save(EXCEL_FILE)


def read_description(day_number: int) -> str:
    """Читает описание лунного дня из Excel (строка day_number, столбец 1)."""
    ensure_excel_exists()
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        cell = ws.cell(row=day_number, column=1)
        return str(cell.value) if cell.value else "Описание отсутствует"
    except Exception as e:
        return f"Ошибка чтения Excel: {e}"


# ═══════════════════════════════════════════════════════════
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ EPHEM
# ═══════════════════════════════════════════════════════════

def _to_naive_utc(dt: datetime) -> datetime:
    """Преобразует aware datetime в naive UTC (как ожидает ephem)."""
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _from_ephem_date(ed: ephem.Date) -> datetime:
    """Преобразует ephem.Date в aware datetime (UTC)."""
    # ephem.Date.datetime() возвращает naive UTC
    naive = ed.datetime()
    return naive.replace(tzinfo=timezone.utc)


# ═══════════════════════════════════════════════════════════
# ВАРИАНТ 1: АДАПТИВНОЕ ДЕЛЕНИЕ (по умолчанию)
# ═══════════════════════════════════════════════════════════

def _get_moon_day_adaptive(now_utc: datetime):
    """
    Адаптивный метод:
    - Вычисляем точную длительность синодического месяца.
    - Если < 29.5 дней → 29 лунных дней, иначе → 30.
    - Делим реальный месяц на N равных частей.

    Returns: (day_number, start_dt_local, end_dt_local, total_days, synodic_days)
    """
    # Точные моменты новолуний (UTC)
    prev_new = _from_ephem_date(ephem.previous_new_moon(_to_naive_utc(now_utc)))
    next_new = _from_ephem_date(ephem.next_new_moon(_to_naive_utc(now_utc)))

    # Реальная длительность синодического месяца в секундах и днях
    synodic_seconds = (next_new - prev_new).total_seconds()
    synodic_days = synodic_seconds / 86400.0

    # Определяем количество лунных дней
    total_lunar_days = 29 if synodic_days < SYNODIC_THRESHOLD_DAYS else 30

    # Длительность одного лунного дня
    lunar_day_seconds = synodic_seconds / total_lunar_days

    # Сколько прошло от начала лунного месяца
    elapsed = (now_utc - prev_new).total_seconds()

    # Номер текущего лунного дня (1-based)
    day_number = int(elapsed / lunar_day_seconds) + 1
    if day_number > total_lunar_days:
        day_number = total_lunar_days
    if day_number < 1:
        day_number = 1

    # Начало и конец текущего лунного дня
    start_utc = prev_new + timedelta(seconds=(day_number - 1) * lunar_day_seconds)
    end_utc = prev_new + timedelta(seconds=day_number * lunar_day_seconds)

    # Конвертируем в локальное время
    local_tz = datetime.now().astimezone().tzinfo
    start_local = start_utc.astimezone(local_tz)
    end_local = end_utc.astimezone(local_tz)

    return day_number, start_local, end_local, total_lunar_days, synodic_days


# ═══════════════════════════════════════════════════════════
# ВАРИАНТ 2: ТИТХИ (ведическая астрология)
# ═══════════════════════════════════════════════════════════

def _get_moon_longitude(dt_utc: datetime) -> float:
    """Возвращает эклиптическую долготу Луны в градусах [0, 360)."""
    observer = ephem.Observer()
    observer.lat = '0'
    observer.lon = '0'
    observer.date = _to_naive_utc(dt_utc)
    moon = ephem.Moon(observer)
    return float(moon.hlon) * 180.0 / 3.141592653589793  # радианы → градусы


def _get_sun_longitude(dt_utc: datetime) -> float:
    """Возвращает эклиптическую долготу Солнца в градусах [0, 360)."""
    observer = ephem.Observer()
    observer.lat = '0'
    observer.lon = '0'
    observer.date = _to_naive_utc(dt_utc)
    sun = ephem.Sun(observer)
    return float(sun.hlon) * 180.0 / 3.141592653589793


def _get_tithi_angle(dt_utc: datetime) -> float:
    """
    Возвращает угол титхи: разность долгот Луны и Солнца в градусах [0, 360).
    0° = новолуние, 180° = полнолуние.
    """
    moon_lon = _get_moon_longitude(dt_utc)
    sun_lon = _get_sun_longitude(dt_utc)
    diff = moon_lon - sun_lon
    while diff < 0:
        diff += 360.0
    while diff >= 360.0:
        diff -= 360.0
    return diff


def _find_tithi_boundary(start_utc: datetime, end_utc: datetime, target_angle: float,
                         max_iter: int = 50, tol_seconds: float = 1.0) -> datetime:
    """
    Бинарный поиск момента, когда угол титхи = target_angle (градусы).
    """
    lo = start_utc
    hi = end_utc
    for _ in range(max_iter):
        mid = lo + (hi - lo) / 2
        angle = _get_tithi_angle(mid)
        # Нормализуем разность к [-180, 180] для корректного сравнения
        diff = ((angle - target_angle + 180) % 360) - 180
        if abs(diff) < 0.001:  # ~0.001° ≈ 0.24 секунды
            break
        if diff < 0:
            lo = mid
        else:
            hi = mid
        if (hi - lo).total_seconds() < tol_seconds:
            break
    return lo + (hi - lo) / 2


def _get_moon_day_tithi(now_utc: datetime):
    """
    Метод титхи:
    - Лунный день = интервал, в котором угол Луна–Солнце растёт на 12°.
    - Всегда 30 титхи в месяце, но длительность варьируется (19–26 ч).
    - Номер титхи = floor(угол / 12°) + 1.

    Returns: (day_number, start_dt_local, end_dt_local, total_days=30, synodic_days)
    """
    # Текущий угол
    angle = _get_tithi_angle(now_utc)
    day_number = int(angle / 12.0) + 1
    if day_number > 30:
        day_number = 30
    if day_number < 1:
        day_number = 1

    # Границы углов для текущей титхи
    start_angle = (day_number - 1) * 12.0
    end_angle = day_number * 12.0

    # Ищем точные моменты границ
    # Окно поиска: ±2 дня от now_utc (титхи не длиннее ~26 ч)
    search_start = now_utc - timedelta(days=2)
    search_end = now_utc + timedelta(days=2)

    tithi_start = _find_tithi_boundary(search_start, now_utc, start_angle)
    tithi_end = _find_tithi_boundary(now_utc, search_end, end_angle)

    # Локальное время
    local_tz = datetime.now().astimezone().tzinfo
    start_local = tithi_start.astimezone(local_tz)
    end_local = tithi_end.astimezone(local_tz)

    # Длительность синодического месяца для справки
    prev_new = _from_ephem_date(ephem.previous_new_moon(_to_naive_utc(now_utc)))
    next_new = _from_ephem_date(ephem.next_new_moon(_to_naive_utc(now_utc)))
    synodic_days = (next_new - prev_new).total_seconds() / 86400.0

    return day_number, start_local, end_local, 30, synodic_days


# ═══════════════════════════════════════════════════════════
# ОБЩИЙ ИНТЕРФЕЙС
# ═══════════════════════════════════════════════════════════

def get_moon_day_info():
    """
    Вычисляет номер текущего лунного дня и границы именно ЭТОГО дня.

    Returns:
        (day_number, start_dt_local, end_dt_local, total_days, synodic_days)
    """
    now_utc = datetime.now(timezone.utc)

    if MOON_DAY_METHOD == "tithi":
        return _get_moon_day_tithi(now_utc)
    else:
        return _get_moon_day_adaptive(now_utc)


def get_moon_phase_name(day_number: int, total_days: int) -> str:
    """
    Возвращает название фазы Луны на основе номера лунного дня.
    """
    ratio = day_number / total_days
    if day_number == 1:
        return "Новолуние"
    elif ratio < 0.25:
        return "Растущая Луна (первая четверть приближается)"
    elif ratio < 0.27:
        return "Первая четверть"
    elif ratio < 0.48:
        return "Растущая Луна (полнолуние приближается)"
    elif ratio < 0.52:
        return "Полнолуние"
    elif ratio < 0.75:
        return "Убывающая Луна (последняя четверть приближается)"
    elif ratio < 0.77:
        return "Последняя четверть"
    else:
        return "Убывающая Луна (новолуние приближается)"


# ═══════════════════════════════════════════════════════════
# ОСНОВНАЯ КОМАНДА
# ═══════════════════════════════════════════════════════════

async def execute(ws_manager, client_ip: str, data: dict,
                  ollama_chat, ollama_model: str, tts_engine, logger):
    """
    Выполняет команду "Лунный день".
    Мгновенная команда — не требует голосового ввода.
    """
    try:
        logger.info("[moon_day] Вычисление лунного дня...")

        day_num, start_dt, end_dt, total_days, synodic_days = get_moon_day_info()
        phase_name = get_moon_phase_name(day_num, total_days)

        # Формат: MM_DD hh:mm
        start_str = start_dt.strftime("%m_%d %H:%M")
        end_str = end_dt.strftime("%m_%d %H:%M")

        # Дополнительная информация для логов
        method_name = "Титхи" if MOON_DAY_METHOD == "tithi" else "Адаптивное деление"
        synodic_str = f"{synodic_days:.2f}"

        log_msg = (
            f"{day_num} лунный день из {total_days}. "
            f"Начало: {start_str}, конец: {end_str}. "
            f"Фаза: {phase_name}. "
            f"Синодический месяц: {synodic_str} дн. "
            f"Метод: {method_name}"
        )
        logger.info(f"[moon_day] {log_msg}")

        # Читаем описание из Excel
        description = read_description(day_num)
        logger.info(f"[moon_day] Описание: {description}")

        # Полный текст для клиента и TTS
        full_text = f"{day_num} лунный день из {total_days}. {phase_name}. Начало: {start_str}, конец: {end_str}. {description}"

        # Отправляем клиенту для отображения в логах
        await ws_manager.send_to(client_ip, {
            "type": "response",
            "text": full_text,
            "module": "command_user",
            "command": "moon_day"
        })

        # Озвучиваем через Silero TTS
        if tts_engine:
            tts_engine.speak(full_text)
            logger.info("[moon_day] Текст озвучен")
        else:
            logger.warning("[moon_day] TTS не инициализирован")

    except Exception as e:
        logger.error(f"[moon_day] Ошибка: {e}")
        await ws_manager.send_to(client_ip, {
            "type": "error",
            "message": f"Ошибка вычисления лунного дня: {e}",
            "module": "command_user",
            "command": "moon_day"
        })